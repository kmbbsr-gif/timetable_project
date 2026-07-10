from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
import traceback

from apps.schools.models import School
from apps.academic.models import AcademicSession, Class, Section, PeriodDefinition, Subject
from apps.teachers.models import Teacher
from .models import TimetableEntry, SubstituteAssignment
from .scheduling.engine import TimetableSolver

import csv
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse

# ---------- Generator ----------
@login_required
def generate_timetable(request):
    if request.method == 'POST':
        school_id = request.POST.get('school_id')
        session_id = request.POST.get('session_id')
        if not school_id or not session_id:
            messages.error(request, "Please select a school and session.")
            return render(request, 'timetable/generate_form.html', {
                'schools': School.objects.all(),
                'sessions': AcademicSession.objects.filter(school__in=School.objects.all()),
            })
        try:
            solver = TimetableSolver(int(school_id), int(session_id))
            count = solver.solve(use_genetic=False)
            if count:
                messages.success(request, f"✅ Timetable generated! {count} entries created.")
                return redirect('dashboard')
            else:
                messages.error(request, "❌ No solution found. Check constraints.")
                return render(request, 'timetable/generate_form.html', {
                    'schools': School.objects.all(),
                    'sessions': AcademicSession.objects.filter(school__in=School.objects.all()),
                })
        except Exception as e:
            messages.error(request, f"⚠️ Error: {str(e)}")
            traceback.print_exc()
            return render(request, 'timetable/generate_form.html', {
                'schools': School.objects.all(),
                'sessions': AcademicSession.objects.filter(school__in=School.objects.all()),
            })

    schools = School.objects.all()
    sessions = AcademicSession.objects.filter(school__in=schools)
    return render(request, 'timetable/generate_form.html', {'schools': schools, 'sessions': sessions})

# ---------- Timetable Index ----------
def timetable_index(request):
    all_classes = Class.objects.filter(is_active=True).order_by('display_order')
    class_sections = []
    sections_data_json = []
    for cls in all_classes:
        sections = Section.objects.filter(class_instance=cls, is_active=True)
        class_sections.append({'class': cls, 'sections': sections})
        sections_data_json.append({
            'class': {'id': cls.id, 'name': cls.name},
            'sections': [{'id': s.id, 'name': s.name} for s in sections]
        })
    classes_with_entries = Class.objects.filter(
        id__in=TimetableEntry.objects.values_list('class_instance_id', flat=True).distinct()
    )
    data = []
    for cls in classes_with_entries:
        sections = Section.objects.filter(
            class_instance=cls,
            id__in=TimetableEntry.objects.filter(class_instance=cls).values_list('section_id', flat=True).distinct()
        )
        data.append({'class': cls, 'sections': sections})
    context = {
        'class_sections': class_sections,
        'sections_data_json': sections_data_json,
        'data': data,
    }
    return render(request, 'timetable/index.html', context)

def select_class(request):
    if request.method == 'GET':
        class_id = request.GET.get('class_id')
        section_id = request.GET.get('section_id')
        if class_id and section_id:
            try:
                class_id = int(class_id)
                section_id = int(section_id)
                return redirect('view_class_timetable', class_id=class_id, section_id=section_id)
            except (ValueError, TypeError):
                pass
    return redirect('timetable_index')

# ---------- Class Timetable ----------
def view_class_timetable(request, class_id, section_id):
    date_str = request.GET.get('date')
    date = None
    if date_str:
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    class_obj = get_object_or_404(Class, id=class_id)
    section = get_object_or_404(Section, id=section_id)

    entries = TimetableEntry.objects.filter(
        class_instance=class_obj,
        section=section
    ).select_related('teacher', 'subject')

    # Determine which days to show
    all_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    if date is not None:
        # Only show the day corresponding to the date
        day_idx = date.weekday()  # Monday=0
        if day_idx >= 6:  # Sunday
            day_idx = 0  # fallback to Monday if Sunday (no school)
        days_to_show = [(day_idx, all_days[day_idx])]
    else:
        days_to_show = [(i, name) for i, name in enumerate(all_days)]

    # Build grid and table as before, but using days_to_show
    grid = {}
    for entry in entries:
        key = f"{entry.day_of_week}_{entry.period_number}"
        grid[key] = entry

    periods = PeriodDefinition.objects.filter(
        school=class_obj.school,
        wing=section.wing
    ).order_by('period_number')
    if not periods:
        periods = PeriodDefinition.objects.filter(school=class_obj.school).order_by('period_number')
    if not periods:
        periods = [{'period_number': i, 'start_time': None, 'end_time': None} for i in range(1, 11)]

    table = []
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = {'period': p, 'cells': []}
        for day_idx, _ in days_to_show:
            key = f"{day_idx}_{pn}"
            row['cells'].append(grid.get(key))
        table.append(row)

    context = {
        'class': class_obj,
        'section': section,
        'table': table,
        'days_with_indices': days_to_show,  # only the days being shown
        'selected_date': date_str,
    }
    return render(request, 'timetable/class_view.html', context)

# ---------- Teacher Timetable ----------
def teacher_timetable_index(request):
    teachers = Teacher.objects.filter(
        id__in=TimetableEntry.objects.values_list('teacher_id', flat=True).distinct()
    ).order_by('name')
    return render(request, 'timetable/teacher_index.html', {'teachers': teachers})

from datetime import datetime

def view_teacher_timetable(request, teacher_id):
    date_str = request.GET.get('date')
    date = None
    if date_str:
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    teacher = get_object_or_404(Teacher, id=teacher_id)
    entries = TimetableEntry.objects.filter(teacher=teacher).select_related('class_instance', 'section', 'subject')

    all_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    if date is not None:
        day_idx = date.weekday()
        if day_idx >= 6:
            day_idx = 0
        days_to_show = [(day_idx, all_days[day_idx])]
    else:
        days_to_show = [(i, name) for i, name in enumerate(all_days)]

    grid = {}
    period_numbers = set()
    for entry in entries:
        if date:
            # Check substitution (same logic as before)
            try:
                sub = SubstituteAssignment.objects.get(
                    absent_teacher=entry.teacher,
                    date=date,
                    period_number=entry.period_number,
                    class_instance=entry.class_instance,
                    section=entry.section
                )
                entry.substitute_teacher = sub.substitute_teacher
            except SubstituteAssignment.DoesNotExist:
                entry.substitute_teacher = None
        else:
            entry.substitute_teacher = None

        key = f"{entry.day_of_week}_{entry.period_number}"
        grid[key] = entry
        period_numbers.add(entry.period_number)

    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=teacher.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})

    table = []
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = {'period': p, 'cells': []}
        for day_idx, _ in days_to_show:
            key = f"{day_idx}_{pn}"
            row['cells'].append(grid.get(key))
        table.append(row)

    context = {
        'teacher': teacher,
        'table': table,
        'days_with_indices': days_to_show,
        'selected_date': date_str,
    }
    return render(request, 'timetable/teacher_view.html', context)
# ---------- Subject Timetable ----------
def subject_timetable_index(request):
    subjects = Subject.objects.filter(
        id__in=TimetableEntry.objects.values_list('subject_id', flat=True).distinct()
    ).order_by('name')
    return render(request, 'timetable/subject_index.html', {'subjects': subjects})

def view_subject_timetable(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    entries = TimetableEntry.objects.filter(subject=subject).select_related('class_instance', 'section', 'teacher')

    # Get all periods for the school and deduplicate by period_number
    all_periods = PeriodDefinition.objects.filter(school=subject.school).order_by('period_number')
    periods = {}
    for p in all_periods:
        if p.period_number not in periods:
            periods[p.period_number] = p
    if not periods:
        for i in range(1, 11):
            periods[i] = {'period_number': i, 'start_time': None, 'end_time': None}
    periods = [periods[pn] for pn in sorted(periods.keys())]

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]

    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        if key not in grid:
            grid[key] = []
        grid[key].append(e)

    table = []
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = {'period': p, 'cells': []}
        for day_idx, _ in days_with_indices:
            key = f"{day_idx}_{pn}"
            row['cells'].append(grid.get(key))
        table.append(row)

    context = {
        'subject': subject,
        'table': table,
        'days_with_indices': days_with_indices,
    }
    return render(request, 'timetable/subject_view.html', context)
# ---------- Master Timetable ----------
def master_timetable(request):
    school = School.objects.first()
    if not school:
        return render(request, 'timetable/master.html', {'data': [], 'days_with_indices': []})

    classes = Class.objects.filter(school=school, is_active=True).order_by('display_order')
    period_numbers = PeriodDefinition.objects.filter(school=school).values_list('period_number', flat=True).distinct()
    if not period_numbers:
        period_numbers = range(1, 11)

    periods = []
    for pn in sorted(set(period_numbers)):
        p_def = PeriodDefinition.objects.filter(school=school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]

    data = []
    for cls in classes:
        sections = Section.objects.filter(class_instance=cls, is_active=True)
        for section in sections:
            entries = TimetableEntry.objects.filter(
                class_instance=cls,
                section=section
            ).select_related('subject', 'teacher')
            grid = {}
            for e in entries:
                key = f"{e.day_of_week}_{e.period_number}"
                grid[key] = e
            table = []
            for p in periods:
                pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
                row = {'period': p, 'cells': []}
                for day_idx, _ in days_with_indices:
                    key = f"{day_idx}_{pn}"
                    row['cells'].append(grid.get(key))
                table.append(row)
            data.append({
                'class': cls,
                'section': section,
                'table': table,
            })

    context = {
        'data': data,
        'days_with_indices': days_with_indices,
    }
    return render(request, 'timetable/master.html', context)

# ---------- Free Periods ----------
def free_periods(request):
    school = School.objects.first()
    if not school:
        return render(request, 'timetable/free_periods.html', {'days_with_indices': [], 'table': [], 'teachers': []})

    # Get unique period numbers from PeriodDefinition (all wings, unique)
    all_period_numbers = PeriodDefinition.objects.filter(school=school).values_list('period_number', flat=True).distinct()
    if not all_period_numbers:
        all_period_numbers = range(1, 11)

    periods = []
    for pn in sorted(set(all_period_numbers)):
        p_def = PeriodDefinition.objects.filter(school=school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]
    teachers = Teacher.objects.filter(is_active=True, school=school)

    # Build occupied set for all teachers
    occupied = set()
    for t_id, day, period in TimetableEntry.objects.all().values_list('teacher_id', 'day_of_week', 'period_number'):
        occupied.add((t_id, day, period))

    # Build table: for each period and day, list free teachers
    table = []
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = {'period': p, 'cells': []}
        for day_idx, _ in days_with_indices:
            free_list = []
            for t in teachers:
                if (t.id, day_idx, pn) not in occupied:
                    free_list.append(t.name)
            row['cells'].append(free_list)
        table.append(row)

    # Teacher-specific lookup (handles date if needed)
    selected_teacher_id = request.GET.get('teacher_id')
    selected_day = request.GET.get('day')
    teacher_free_info = None
    selected_teacher = None

    if selected_teacher_id:
        try:
            selected_teacher = Teacher.objects.get(id=int(selected_teacher_id), school=school)
            occupied_teacher = set()
            for entry in TimetableEntry.objects.filter(teacher=selected_teacher):
                occupied_teacher.add((entry.day_of_week, entry.period_number))

            if selected_day == 'all' or selected_day is None:
                free_by_day = {}
                for day_idx, day_name in days_with_indices:
                    free_periods_list = []
                    for p in periods:
                        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
                        if (day_idx, pn) not in occupied_teacher:
                            free_periods_list.append(pn)
                    free_by_day[day_name] = free_periods_list
                teacher_free_info = {'type': 'all_days', 'data': free_by_day}
            else:
                day_idx = int(selected_day)
                day_name = days[day_idx]
                free_periods_list = []
                for p in periods:
                    pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
                    if (day_idx, pn) not in occupied_teacher:
                        free_periods_list.append(pn)
                teacher_free_info = {'type': 'single_day', 'day': day_name, 'periods': free_periods_list}
        except (ValueError, Teacher.DoesNotExist):
            pass

    context = {
        'days_with_indices': days_with_indices,
        'table': table,
        'teachers': teachers,
        'selected_teacher': selected_teacher,
        'selected_day': selected_day,
        'teacher_free_info': teacher_free_info,
        'periods': periods,
    }
    return render(request, 'timetable/free_periods.html', context)

# ---------- Debug ----------
def debug_timetable(request):
    entries = TimetableEntry.objects.all().select_related('class_instance', 'section', 'subject', 'teacher')
    output = []
    for e in entries:
        output.append(
            f"Class: {e.class_instance.id} - {e.class_instance.name}, "
            f"Section: {e.section.id} - {e.section.name}, "
            f"Day: {e.day_of_week}, "
            f"Period: {e.period_number}, "
            f"Subject: {e.subject.name}, "
            f"Teacher: {e.teacher.name}"
        )
    return render(request, 'timetable/debug.html', {'entries': output})

# ---------- Teacher Absence Management ----------
def manage_absence(request):
    """
    Manage teacher absence: show schedule and allow substitute assignment.
    """
    # Get parameters: teacher_id, date
    teacher_id = request.GET.get('teacher_id')
    date_str = request.GET.get('date')
    selected_date = None
    day_of_week = None
    teacher = None
    schedule = []
    free_teachers_by_period = {}
    substitutions = {}
    
    if teacher_id and date_str:
        try:
            teacher = get_object_or_404(Teacher, id=int(teacher_id))
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            day_of_week = selected_date.weekday()  # Monday=0
            
            # Get periods for that day (assuming periods are defined for the school)
            periods = PeriodDefinition.objects.filter(school=teacher.school).order_by('period_number')
            # But we need periods that are active on this day? We'll use all periods defined.
            # Better: get periods that are used in the timetable entries.
            period_numbers = TimetableEntry.objects.filter(
                teacher=teacher,
                day_of_week=day_of_week
            ).values_list('period_number', flat=True).distinct()
            periods = PeriodDefinition.objects.filter(
                school=teacher.school,
                period_number__in=period_numbers
            ).order_by('period_number')
            
            # Get the teacher's timetable for this day
            entries = TimetableEntry.objects.filter(
                teacher=teacher,
                day_of_week=day_of_week
            ).select_related('class_instance', 'section', 'subject')
            
            # Build schedule list
            for entry in entries:
                schedule.append({
                    'period': entry.period_number,
                    'class_instance': entry.class_instance,
                    'section': entry.section,
                    'subject': entry.subject,
                    'entry': entry,
                })
            
            # For each period, find free teachers (excluding the absent teacher)
            all_teachers = Teacher.objects.filter(is_active=True, school=teacher.school).exclude(id=teacher.id)
            # Build occupied set for this day
            occupied = {}
            for t in all_teachers:
                occupied[t.id] = set()
            # Populate occupied periods for each teacher
            for t in all_teachers:
                occupied_periods = TimetableEntry.objects.filter(
                    teacher=t,
                    day_of_week=day_of_week
                ).values_list('period_number', flat=True)
                occupied[t.id] = set(occupied_periods)
            
            free_teachers_by_period = {}
            # Also get subject eligibility? For now, just list all free teachers.
            for entry in schedule:
                pn = entry['period']
                 
                free_list = []
                for t in all_teachers:
                    if pn not in occupied[t.id]:
                        free_list.append(t)
                free_teachers_by_period[pn] = free_list
                
                for entry in schedule:
                    pn = entry['period']
                    entry['free_teachers'] = free_teachers_by_period.get(pn, [])
            
            # Get existing substitutions for this teacher on this date
            sub_assignments = SubstituteAssignment.objects.filter(
                absent_teacher=teacher,
                date=selected_date
            ).select_related('substitute_teacher', 'class_instance', 'section', 'subject')
            for sub in sub_assignments:
                substitutions[(sub.period_number, sub.class_instance.id, sub.section.id)] = sub.substitute_teacher
            
        except ValueError:
            messages.error(request, "Invalid data.")
            return redirect('timetable_index')
    
    # Handle POST: save substitutions
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher_id')
        date_str = request.POST.get('date')
        if not teacher_id or not date_str:
            messages.error(request, "Missing teacher or date.")
            return redirect('manage_absence')
        try:
            teacher = get_object_or_404(Teacher, id=int(teacher_id))
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            # Process each period
            for key, value in request.POST.items():
                if key.startswith('sub_'):
                    # key format: sub_<period>_<class_id>_<section_id>
                    parts = key.split('_')
                    if len(parts) == 4:
                        period = int(parts[1])
                        class_id = int(parts[2])
                        section_id = int(parts[3])
                        substitute_id = int(value) if value else None
                        # Get the original timetable entry to get subject
                        orig_entry = TimetableEntry.objects.filter(
                            teacher=teacher,
                            day_of_week=selected_date.weekday(),
                            period_number=period,
                            class_instance_id=class_id,
                            section_id=section_id
                        ).first()
                        if not orig_entry:
                            continue
                        # Delete existing substitution if any
                        SubstituteAssignment.objects.filter(
                            absent_teacher=teacher,
                            date=selected_date,
                            period_number=period,
                            class_instance_id=class_id,
                            section_id=section_id
                        ).delete()
                        if substitute_id:
                            substitute_teacher = get_object_or_404(Teacher, id=substitute_id)
                            SubstituteAssignment.objects.create(
                                absent_teacher=teacher,
                                substitute_teacher=substitute_teacher,
                                date=selected_date,
                                period_number=period,
                                class_instance=orig_entry.class_instance,
                                section=orig_entry.section,
                                subject=orig_entry.subject,
                            )
            messages.success(request, "Substitutions saved successfully.")
            # Redirect back with same parameters
            return redirect(f"{request.path}?teacher_id={teacher_id}&date={date_str}")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
    
    # GET: prepare form data
    teachers = Teacher.objects.filter(is_active=True).order_by('name')
    context = {
        'teachers': teachers,
        'selected_teacher': teacher,
        'selected_date': selected_date,
        'schedule': schedule,
        'free_teachers_by_period': free_teachers_by_period,
        'substitutions': substitutions,
        'periods': [entry['period'] for entry in schedule],
    }
    return render(request, 'timetable/manage_absence.html', context)


# ---------- Teacher Load API ----------
def teacher_load(request, teacher_id):
    try:
        teacher = get_object_or_404(Teacher, id=teacher_id)
        load = teacher.total_assigned_periods()
        remaining = teacher.remaining_periods()
        return JsonResponse({
            'load': load,
            'remaining': remaining,
            'max': teacher.max_weekly_load,
        })
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
    
def get_conflict_suggestions(source_entry, target_entry=None, target_day=None, target_period=None):
    school = source_entry.school
    if target_entry:
        target_day = target_entry.day_of_week
        target_period = target_entry.period_number
    else:
        target_day = int(target_day) if target_day is not None else None
        target_period = int(target_period) if target_period is not None else None

    suggestions = {
        'free_teachers': [],
        'free_periods': [],
    }

    if target_day is not None and target_period is not None:
        # Free teachers for that day/period (excluding the source teacher)
        all_teachers = Teacher.objects.filter(school=school, is_active=True).exclude(id=source_entry.teacher_id)
        occupied_teachers = TimetableEntry.objects.filter(
            school=school,
            day_of_week=target_day,
            period_number=target_period
        ).values_list('teacher_id', flat=True)
        free_teachers = all_teachers.exclude(id__in=occupied_teachers)
        suggestions['free_teachers'] = list(free_teachers.values('id', 'name')[:10])

        # Free periods for the source class (same class/section, other days/periods)
        source_class = source_entry.class_instance
        source_section = source_entry.section
        occupied_slots = TimetableEntry.objects.filter(
            class_instance=source_class,
            section=source_section
        ).values_list('day_of_week', 'period_number')
        occupied_set = set(occupied_slots)
        periods = PeriodDefinition.objects.filter(school=school).values_list('period_number', flat=True).distinct()
        days = range(6)
        free_slots = []
        for day in days:
            for pn in periods:
                if (day, pn) not in occupied_set:
                    free_slots.append({'day': day, 'period': pn})
        suggestions['free_periods'] = free_slots[:10]

    return suggestions

def get_temp_slot(school, session, excluded_ids):
    """Find a (day, period) that is not occupied by any timetable entry."""
    occupied = set(
        TimetableEntry.objects.filter(school=school, session=session)
        .exclude(id__in=excluded_ids)
        .values_list('day_of_week', 'period_number')
    )
    # Try Sunday (6) with high period numbers (99, 98, ...)
    # day 6 is Sunday, period 99 is unlikely to be used
    for day in [6, 5, 4, 3, 2, 1, 0]:
        for period in [99, 98, 97, 96]:
            if (day, period) not in occupied:
                return day, period
    # Fallback: find any free slot by scanning all days and periods 1..20
    for day in range(7):
        for period in range(1, 21):
            if (day, period) not in occupied:
                return day, period
    raise Exception("No free temporary slot found")

# ---------- Swap Entries (Drag & Drop) ----------
from django.db import transaction

@login_required
@csrf_exempt
def swap_entries(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    source_id = request.POST.get('source_id')
    target_id = request.POST.get('target_id')
    move = request.POST.get('move') == 'true'
    override = request.POST.get('override') == 'true'
    swap_teachers = request.POST.get('swap_teachers') == 'true'
    assign_free_teacher = request.POST.get('assign_free_teacher')

    if not source_id:
        return JsonResponse({'error': 'Missing source entry'}, status=400)

    try:
        source = TimetableEntry.objects.get(id=source_id)
    except TimetableEntry.DoesNotExist:
        return JsonResponse({'error': 'Source entry not found'}, status=404)

    conflicts = []

    # ---- Handle assigning a free teacher ----
    if assign_free_teacher:
        try:
            new_teacher = Teacher.objects.get(id=int(assign_free_teacher))
            # Check if new teacher is free at that day/period
            # (we already checked free teachers, so just assign)
            source.teacher = new_teacher
            source.save()
            return JsonResponse({'success': True})
        except Teacher.DoesNotExist:
            return JsonResponse({'error': 'Teacher not found'}, status=404)

    # ---- Swap Teachers (exchange teacher fields) ----
    if swap_teachers and target_id:
        try:
            target = TimetableEntry.objects.get(id=target_id)
        except TimetableEntry.DoesNotExist:
            return JsonResponse({'error': 'Target entry not found'}, status=404)
        source_teacher = source.teacher
        target_teacher = target.teacher
        source.teacher = target_teacher
        target.teacher = source_teacher
        with transaction.atomic():
            source.save()
            target.save()
        return JsonResponse({'success': True})

    # ---- Move to empty cell (no target entry) ----
    if move:
        target_class_id = request.POST.get('target_class')
        target_section_id = request.POST.get('target_section')
        target_day = request.POST.get('target_day')
        target_period = request.POST.get('target_period')

        if not all([target_class_id, target_section_id, target_day, target_period]):
            return JsonResponse({'error': 'Missing target position'}, status=400)

        # Check if target cell is empty
        existing = TimetableEntry.objects.filter(
            class_instance_id=target_class_id,
            section_id=target_section_id,
            day_of_week=target_day,
            period_number=target_period
        ).exists()
        if existing:
            return JsonResponse({'error': 'Target cell is not empty'}, status=400)

        # Check teacher conflict
        overlap_teacher = TimetableEntry.objects.filter(
            teacher=source.teacher,
            day_of_week=target_day,
            period_number=target_period
        ).exclude(id=source.id)
        if overlap_teacher.exists():
            conflicts.append(f"Teacher {source.teacher.name} already has a class at that time")

        if conflicts and not override:
            # Generate suggestions
            suggestions = get_conflict_suggestions(source, target_day=int(target_day), target_period=int(target_period))
            return JsonResponse({
                'conflicts': conflicts,
                'requires_override': True,
                'suggestions': suggestions,
            }, status=409)

        # Perform move
        with transaction.atomic():
            source.day_of_week = int(target_day)
            source.period_number = int(target_period)
            source.save()
        return JsonResponse({'success': True})

    # ---- Swap positions (both source and target exist) ----
        # ---- Swap positions (both source and target exist) ----
    if not move and target_id:
        try:
            target = TimetableEntry.objects.get(id=target_id)
        except TimetableEntry.DoesNotExist:
            return JsonResponse({'error': 'Target entry not found'}, status=404)

        # Check conflicts (teacher/class overlaps) – keep your existing conflict checks here
        # ... (same conflict checks as before) ...

        if conflicts and not override:
            suggestions = get_conflict_suggestions(source, target)
            return JsonResponse({
                'conflicts': conflicts,
                'requires_override': True,
                'suggestions': suggestions,
            }, status=409)

        # Perform swap using a temporary free slot
        with transaction.atomic():
            src_day, src_period = source.day_of_week, source.period_number
            tgt_day, tgt_period = target.day_of_week, target.period_number

            # Get a free temporary slot
            temp_day, temp_period = get_temp_slot(source.school, source.session, [source.id, target.id])

            # Move source to temp slot
            source.day_of_week = temp_day
            source.period_number = temp_period
            source.save()

            # Move target to source's old position
            target.day_of_week = src_day
            target.period_number = src_period
            target.save()

            # Move source to target's old position
            source.day_of_week = tgt_day
            source.period_number = tgt_period
            source.save()

        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Invalid operation'}, status=400)
def export_pdf(request):
    # Get master timetable data
    school = School.objects.first()
    if not school:
        return HttpResponse("No school found.", status=404)
    classes = Class.objects.filter(school=school, is_active=True).order_by('display_order')
    period_numbers = PeriodDefinition.objects.filter(school=school).values_list('period_number', flat=True).distinct()
    if not period_numbers:
        period_numbers = range(1, 11)
    periods = []
    for pn in sorted(set(period_numbers)):
        p_def = PeriodDefinition.objects.filter(school=school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]

    # Build data structure
    rows = []
    header = ['Class', 'Section', 'Period'] + days
    rows.append(header)
    for cls in classes:
        sections = Section.objects.filter(class_instance=cls, is_active=True)
        for section in sections:
            entries = TimetableEntry.objects.filter(
                class_instance=cls,
                section=section
            ).select_related('subject', 'teacher')
            grid = {}
            for e in entries:
                key = f"{e.day_of_week}_{e.period_number}"
                grid[key] = e
            for p in periods:
                pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
                row = [cls.name, section.name, f"P{pn}"]
                for day_idx, _ in days_with_indices:
                    key = f"{day_idx}_{pn}"
                    entry = grid.get(key)
                    if entry:
                        row.append(f"{entry.subject.code}\n{entry.teacher.name}")
                    else:
                        row.append("")
                rows.append(row)

    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="master_timetable.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    # Title
    title = Paragraph("Master Timetable", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.25*inch))

    # Table
    table = Table(rows, repeatRows=2)  # repeat header row
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 1), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 1), 10),
        ('FONTSIZE', (0, 2), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('SPAN', (0, 0), (0, 1)),  # merge Class header
        ('SPAN', (1, 0), (1, 1)),  # merge Section header
        ('SPAN', (2, 0), (2, 1)),  # merge Period header
    ]))
    elements.append(table)

    doc.build(elements)
    return response

def export_excel(request):
    # Build data (same as PDF)
    school = School.objects.first()
    if not school:
        return HttpResponse("No school found.", status=404)
    classes = Class.objects.filter(school=school, is_active=True).order_by('display_order')
    period_numbers = PeriodDefinition.objects.filter(school=school).values_list('period_number', flat=True).distinct()
    if not period_numbers:
        period_numbers = range(1, 11)
    periods = []
    for pn in sorted(set(period_numbers)):
        p_def = PeriodDefinition.objects.filter(school=school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Timetable"

    # Headers
    headers = ['Class', 'Section', 'Period'] + days
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 2
    for cls in classes:
        sections = Section.objects.filter(class_instance=cls, is_active=True)
        for section in sections:
            entries = TimetableEntry.objects.filter(
                class_instance=cls,
                section=section
            ).select_related('subject', 'teacher')
            grid = {}
            for e in entries:
                key = f"{e.day_of_week}_{e.period_number}"
                grid[key] = e
            for p in periods:
                pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
                row = [cls.name, section.name, f"P{pn}"]
                for day in days:
                    day_idx = days.index(day)
                    key = f"{day_idx}_{pn}"
                    entry = grid.get(key)
                    if entry:
                        row.append(f"{entry.subject.code}\n{entry.teacher.name}")
                    else:
                        row.append("")
                for col, val in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                row_num += 1

    # Adjust column widths
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="master_timetable.xlsx"'
    wb.save(response)
    return response
def export_csv(request):
    school = School.objects.first()
    if not school:
        return HttpResponse("No school found.", status=404)
    classes = Class.objects.filter(school=school, is_active=True).order_by('display_order')
    period_numbers = PeriodDefinition.objects.filter(school=school).values_list('period_number', flat=True).distinct()
    if not period_numbers:
        period_numbers = range(1, 11)
    periods = []
    for pn in sorted(set(period_numbers)):
        p_def = PeriodDefinition.objects.filter(school=school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="master_timetable.csv"'
    writer = csv.writer(response)
    header = ['Class', 'Section', 'Period'] + days
    writer.writerow(header)
    for cls in classes:
        sections = Section.objects.filter(class_instance=cls, is_active=True)
        for section in sections:
            entries = TimetableEntry.objects.filter(
                class_instance=cls,
                section=section
            ).select_related('subject', 'teacher')
            grid = {}
            for e in entries:
                key = f"{e.day_of_week}_{e.period_number}"
                grid[key] = e
            for p in periods:
                pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
                row = [cls.name, section.name, f"P{pn}"]
                for day_idx, _ in enumerate(days):
                    key = f"{day_idx}_{pn}"
                    entry = grid.get(key)
                    if entry:
                        row.append(f"{entry.subject.code} {entry.teacher.name}")
                    else:
                        row.append("")
                writer.writerow(row)
    return response

from django.http import HttpResponse
import csv
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ---- Helper to build table rows for a given queryset ----
def _build_timetable_rows(entries, periods, classes_sections=True):
    # Build grid and rows similar to master
    # but we need to handle class/section or teacher/subject context.
    # We'll build a list of rows with headers.
    pass  # we'll implement in each export

# ---- Class Timetable Export ----
def export_class_pdf(request, class_id, section_id):
    class_obj = get_object_or_404(Class, id=class_id)
    section = get_object_or_404(Section, id=section_id)
    entries = TimetableEntry.objects.filter(class_instance=class_obj, section=section).select_related('subject', 'teacher')
    # Get periods from entries
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=class_obj.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]
    # Build grid
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        grid[key] = e
    # Build rows
    rows = [['Period'] + days]
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day_idx, _ in days_with_indices:
            key = f"{day_idx}_{pn}"
            entry = grid.get(key)
            if entry:
                row.append(f"{entry.subject.code}\n{entry.teacher.name}")
            else:
                row.append("")
        rows.append(row)
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="class_{class_obj.name}_{section.name}_timetable.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    title = Paragraph(f"Timetable: {class_obj.name} - {section.name}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.25*inch))
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

def export_class_excel(request, class_id, section_id):
    class_obj = get_object_or_404(Class, id=class_id)
    section = get_object_or_404(Section, id=section_id)
    entries = TimetableEntry.objects.filter(class_instance=class_obj, section=section).select_related('subject', 'teacher')
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=class_obj.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        grid[key] = e
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_obj.name}_{section.name}"
    headers = ['Period'] + days
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row_num = 2
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day in days:
            day_idx = days.index(day)
            key = f"{day_idx}_{pn}"
            entry = grid.get(key)
            if entry:
                row.append(f"{entry.subject.code} {entry.teacher.name}")
            else:
                row.append("")
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        row_num += 1
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 15
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="class_{class_obj.name}_{section.name}_timetable.xlsx"'
    wb.save(response)
    return response

def export_class_csv(request, class_id, section_id):
    class_obj = get_object_or_404(Class, id=class_id)
    section = get_object_or_404(Section, id=section_id)
    entries = TimetableEntry.objects.filter(class_instance=class_obj, section=section).select_related('subject', 'teacher')
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=class_obj.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        grid[key] = e
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="class_{class_obj.name}_{section.name}_timetable.csv"'
    writer = csv.writer(response)
    writer.writerow(['Period'] + days)
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day_idx in range(6):
            key = f"{day_idx}_{pn}"
            entry = grid.get(key)
            if entry:
                row.append(f"{entry.subject.code} {entry.teacher.name}")
            else:
                row.append("")
        writer.writerow(row)
    return response

# ---- Teacher Timetable Export ----
def export_teacher_pdf(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    entries = TimetableEntry.objects.filter(teacher=teacher).select_related('class_instance', 'section', 'subject')
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=teacher.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        grid[key] = e
    rows = [['Period'] + days]
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day_idx, _ in days_with_indices:
            key = f"{day_idx}_{pn}"
            entry = grid.get(key)
            if entry:
                row.append(f"{entry.class_instance.name} {entry.section.name}\n{entry.subject.code}")
            else:
                row.append("")
        rows.append(row)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="teacher_{teacher.name}_timetable.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    title = Paragraph(f"Timetable: {teacher.name}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.25*inch))
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

def export_teacher_excel(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    entries = TimetableEntry.objects.filter(teacher=teacher).select_related('class_instance', 'section', 'subject')
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=teacher.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        grid[key] = e
    wb = Workbook()
    ws = wb.active
    ws.title = teacher.name[:31]
    headers = ['Period'] + days
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row_num = 2
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day in days:
            day_idx = days.index(day)
            key = f"{day_idx}_{pn}"
            entry = grid.get(key)
            if entry:
                row.append(f"{entry.class_instance.name} {entry.section.name} {entry.subject.code}")
            else:
                row.append("")
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        row_num += 1
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 15
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="teacher_{teacher.name}_timetable.xlsx"'
    wb.save(response)
    return response

def export_teacher_csv(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    entries = TimetableEntry.objects.filter(teacher=teacher).select_related('class_instance', 'section', 'subject')
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=teacher.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        grid[key] = e
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="teacher_{teacher.name}_timetable.csv"'
    writer = csv.writer(response)
    writer.writerow(['Period'] + days)
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day_idx in range(6):
            key = f"{day_idx}_{pn}"
            entry = grid.get(key)
            if entry:
                row.append(f"{entry.class_instance.name} {entry.section.name} {entry.subject.code}")
            else:
                row.append("")
        writer.writerow(row)
    return response

# ---- Subject Timetable Export ----
def export_subject_pdf(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    entries = TimetableEntry.objects.filter(subject=subject).select_related('class_instance', 'section', 'teacher')
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=subject.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        if key not in grid:
            grid[key] = []
        grid[key].append(e)
    rows = [['Period'] + days]
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day_idx, _ in days_with_indices:
            key = f"{day_idx}_{pn}"
            entry_list = grid.get(key)
            if entry_list:
                # Combine multiple entries in one cell
                cell_text = "\n".join([f"{e.class_instance.name} {e.section.name} ({e.teacher.name})" for e in entry_list])
                row.append(cell_text)
            else:
                row.append("")
        rows.append(row)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="subject_{subject.name}_timetable.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    title = Paragraph(f"Timetable: {subject.name}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.25*inch))
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

def export_subject_excel(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    entries = TimetableEntry.objects.filter(subject=subject).select_related('class_instance', 'section', 'teacher')
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=subject.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        if key not in grid:
            grid[key] = []
        grid[key].append(e)
    wb = Workbook()
    ws = wb.active
    ws.title = subject.name[:31]
    headers = ['Period'] + days
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row_num = 2
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day in days:
            day_idx = days.index(day)
            key = f"{day_idx}_{pn}"
            entry_list = grid.get(key)
            if entry_list:
                row.append(", ".join([f"{e.class_instance.name} {e.section.name} ({e.teacher.name})" for e in entry_list]))
            else:
                row.append("")
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        row_num += 1
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 15
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="subject_{subject.name}_timetable.xlsx"'
    wb.save(response)
    return response

def export_subject_csv(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    entries = TimetableEntry.objects.filter(subject=subject).select_related('class_instance', 'section', 'teacher')
    period_numbers = entries.values_list('period_number', flat=True).distinct()
    periods = []
    for pn in sorted(period_numbers):
        p_def = PeriodDefinition.objects.filter(school=subject.school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    grid = {}
    for e in entries:
        key = f"{e.day_of_week}_{e.period_number}"
        if key not in grid:
            grid[key] = []
        grid[key].append(e)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="subject_{subject.name}_timetable.csv"'
    writer = csv.writer(response)
    writer.writerow(['Period'] + days)
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = [f"P{pn}"]
        for day_idx in range(6):
            key = f"{day_idx}_{pn}"
            entry_list = grid.get(key)
            if entry_list:
                row.append(", ".join([f"{e.class_instance.name} {e.section.name}" for e in entry_list]))
            else:
                row.append("")
        writer.writerow(row)
    return response
def free_slots(request):
    school = School.objects.first()
    if not school:
        return render(request, 'timetable/free_slots.html', {'data': []})

    # Get all active classes and sections
    classes = Class.objects.filter(school=school, is_active=True).order_by('display_order')
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]

    # Get all periods for each wing (we'll need to match periods to section wings)
    wing_periods = {}
    for wing in ['Junior', 'Middle', 'Senior']:
        periods = PeriodDefinition.objects.filter(school=school, wing=wing).order_by('period_number')
        if periods:
            wing_periods[wing] = periods
        else:
            # Fallback: use all periods
            periods = PeriodDefinition.objects.filter(school=school).order_by('period_number')
            if periods:
                wing_periods[wing] = periods

    # Build data
    data = []
    for cls in classes:
        sections = Section.objects.filter(class_instance=cls, is_active=True)
        for section in sections:
            wing = section.wing
            periods = wing_periods.get(wing, [])
            if not periods:
                continue

            # Get assigned entries for this class-section
            entries = TimetableEntry.objects.filter(class_instance=cls, section=section)
            assigned = {}
            for e in entries:
                key = (e.day_of_week, e.period_number)
                assigned[key] = e

            # Build table: for each period, mark if assigned
            table = []
            for p in periods:
                pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
                row = {'period': p, 'cells': []}
                for day_idx, _ in days_with_indices:
                    key = (day_idx, pn)
                    entry = assigned.get(key)
                    row['cells'].append(entry)
                table.append(row)

            data.append({
                'class': cls,
                'section': section,
                'table': table,
                'days_with_indices': days_with_indices,
            })

    context = {
        'data': data,
    }
    return render(request, 'timetable/free_slots.html', context)

def free_periods_by_date(request):
    from datetime import datetime
    school = School.objects.first()
    if not school:
        return render(request, 'timetable/free_by_date.html', {'days_with_indices': [], 'table': [], 'teachers': []})

    # Get all period definitions (deduplicated)
    all_period_numbers = PeriodDefinition.objects.filter(school=school).values_list('period_number', flat=True).distinct()
    if not all_period_numbers:
        all_period_numbers = range(1, 11)
    periods = []
    for pn in sorted(set(all_period_numbers)):
        p_def = PeriodDefinition.objects.filter(school=school, period_number=pn).first()
        if p_def:
            periods.append(p_def)
        else:
            periods.append({'period_number': pn, 'start_time': None, 'end_time': None})

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    days_with_indices = [(i, name) for i, name in enumerate(days)]
    teachers = Teacher.objects.filter(is_active=True, school=school)

    # Handle date parameter
    date_str = request.GET.get('date')
    selected_weekday = None
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            selected_weekday = selected_date.weekday()  # Monday=0
        except ValueError:
            pass

    # If date is given, show only that day; otherwise show all days
    if selected_weekday is not None:
        day_indices = [selected_weekday]
        day_names = [days[selected_weekday]]
    else:
        day_indices = [i for i, _ in days_with_indices]
        day_names = [name for _, name in days_with_indices]

    # Build occupied set for all teachers
    occupied = set()
    for t_id, day, period in TimetableEntry.objects.all().values_list('teacher_id', 'day_of_week', 'period_number'):
        occupied.add((t_id, day, period))

    # Build table for the selected days only
    table = []
    for p in periods:
        pn = p.period_number if hasattr(p, 'period_number') else p['period_number']
        row = {'period': p, 'cells': []}
        for day_idx in day_indices:
            free_list = []
            for t in teachers:
                if (t.id, day_idx, pn) not in occupied:
                    free_list.append(t.name)
            row['cells'].append(free_list)
        table.append(row)

    context = {
        'days_with_indices': list(zip(day_indices, day_names)),
        'table': table,
        'teachers': teachers,
        'selected_date': date_str,
    }
    return render(request, 'timetable/free_by_date.html', context)
from .models import SubstituteAssignment

def get_teacher_for_slot(class_id, section_id, day_of_week, period_number, date=None):
    """
    Return the teacher for a given slot.
    If a substitution exists for that date, return the substitute teacher.
    Otherwise, return the original teacher from the timetable entry.
    """
    # Try to find a timetable entry for this slot
    try:
        entry = TimetableEntry.objects.get(
            class_instance_id=class_id,
            section_id=section_id,
            day_of_week=day_of_week,
            period_number=period_number
        )
    except TimetableEntry.DoesNotExist:
        return None

    if date:
        # Check if a substitution exists for this slot on that date
        try:
            sub = SubstituteAssignment.objects.get(
                absent_teacher=entry.teacher,
                date=date,
                period_number=period_number,
                class_instance_id=class_id,
                section_id=section_id
            )
            return sub.substitute_teacher
        except SubstituteAssignment.DoesNotExist:
            pass
    return entry.teacher